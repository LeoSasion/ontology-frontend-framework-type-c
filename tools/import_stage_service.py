from __future__ import annotations

"""Sealed Parquet staging for the million-scale import data plane."""

import json
import os
import shutil
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Iterable, Iterator, Mapping, Sequence
from uuid import uuid4

from bi_cli_core import DB_PATH, now_iso
from dataset_version_store import (
    INTERNAL_ROW_ID,
    INTERNAL_SCHEMA_FIELDS,
    canonical_fingerprint,
    file_sha256,
    workspace_bucket,
)
from workspace_mutation_lock_service import workspace_mutation_lock


IMPORT_STAGE_SCHEMA = "aibi-import-stage/v2"
IMPORT_STAGE_PARSER_VERSION = "duckdb-parquet-v3-single-pass-csv"
DEFAULT_STAGE_TTL_SECONDS = 7 * 24 * 60 * 60
DEFAULT_STAGE_QUOTA_BYTES = 20 * 1024 * 1024 * 1024
MAX_TEST_HELPER_ROWS = 50_000
PARQUET_FILE = "data.parquet"
MANIFEST_FILE = "manifest.json"
CSV_TYPE_SAMPLE_SIZE = 20_480


def _duckdb() -> Any:
    try:
        import duckdb  # type: ignore
    except ImportError as error:  # pragma: no cover - deployment contract
        raise RuntimeError("DuckDB is required for Parquet import staging.") from error
    return duckdb


def _bounded_int(value: str | int | None, fallback: int, minimum: int, maximum: int) -> int:
    try:
        parsed = int(value or fallback)
    except (TypeError, ValueError):
        parsed = fallback
    return max(minimum, min(parsed, maximum))


def _quote_identifier(value: str) -> str:
    return '"' + str(value).replace('"', '""') + '"'


def _sql_literal(value: str | Path) -> str:
    return "'" + str(value).replace("'", "''") + "'"


def configured_import_stage_root(database_path: Path = DB_PATH) -> Path:
    configured = str(os.environ.get("AIBI_IMPORT_STAGE_ROOT") or "").strip()
    return Path(configured).expanduser().resolve() if configured else (database_path.parent / "import-staging-v2").resolve()


def _validate_stage_key(stage_key: str) -> str:
    text = str(stage_key or "")
    suffix = text[6:] if text.startswith("stage_") else ""
    if len(text) != 30 or len(suffix) != 24 or any(character not in "0123456789abcdef" for character in suffix):
        raise ValueError("Import stage key is invalid.")
    return text


def _stage_directory(root: Path, workspace_id: str, stage_key: str) -> Path:
    validated = _validate_stage_key(stage_key)
    directory = (root / "workspaces" / workspace_bucket(workspace_id) / "stages" / validated).resolve()
    if root != directory and root not in directory.parents:
        raise PermissionError("Import stage escaped the configured root.")
    return directory


def _stage_paths(root: Path, workspace_id: str, stage_key: str) -> tuple[Path, Path]:
    directory = _stage_directory(root, workspace_id, stage_key)
    return directory / PARQUET_FILE, directory / MANIFEST_FILE


def _parse_iso(value: str) -> datetime:
    parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
    return parsed if parsed.tzinfo else parsed.replace(tzinfo=timezone.utc)


def _expires_at(created_at: str, ttl_seconds: int) -> str:
    return (_parse_iso(created_at) + timedelta(seconds=ttl_seconds)).astimezone(timezone.utc).isoformat()


def _root_usage(root: Path) -> int:
    if not root.exists():
        return 0
    return sum(
        path.stat().st_size
        for path in root.rglob("*")
        if path.is_file() and not path.is_symlink()
    )


def _fsync_file(path: Path) -> None:
    with path.open("r+b") as stream:
        os.fsync(stream.fileno())


def _fsync_directory(path: Path) -> None:
    if os.name == "nt":
        return
    descriptor = os.open(path, os.O_RDONLY)
    try:
        os.fsync(descriptor)
    finally:
        os.close(descriptor)


def _normalize_headers(headers: Sequence[Any]) -> list[str]:
    normalized: list[str] = []
    counts: dict[str, int] = {}
    for index, raw in enumerate(headers, start=1):
        base = str(raw or "").strip() or f"column_{index}"
        if base.startswith("__aibi_"):
            raise ValueError("Source columns cannot use the reserved __aibi_ prefix.")
        count = counts.get(base, 0) + 1
        counts[base] = count
        normalized.append(base if count == 1 else f"{base}_{count}")
    if not normalized:
        raise ValueError("Import source has no headers.")
    return normalized


def _identifier_field(field: str) -> bool:
    lowered = str(field).strip().casefold()
    return (
        lowered in {"id", "iid", "tid", "oid", "uid", "gid", "sid", "pid", "rid"}
        or lowered.endswith(("_id", "_iid"))
        or any(token in lowered for token in ("code", "sku", "phone", "mobile"))
        or any(token in field for token in ("编号", "编码", "单号", "标识", "手机", "电话", "邮编"))
    )


def _measure_field(field: str) -> bool:
    lowered = str(field).strip().casefold()
    return any(
        token in lowered
        for token in ("amount", "value", "total", "score", "rate", "quantity", "count", "price", "cost", "revenue", "profit")
    ) or any(token in field for token in ("金额", "数值", "合计", "得分", "比例", "数量", "次数", "价格", "收入", "成本", "利润", "费用"))


def _csv_relation(path: Path, forced_varchar: Sequence[str] = ()) -> str:
    forced_types = ""
    if forced_varchar:
        forced_types = ", types = {" + ", ".join(
            f"{_sql_literal(field)}: 'VARCHAR'" for field in forced_varchar
        ) + "}"
    return (
        f"read_csv({_sql_literal(path)}, header = true, auto_detect = true, sample_size = {CSV_TYPE_SAMPLE_SIZE}, "
        "all_varchar = false, ignore_errors = false, null_padding = false, "
        "auto_type_candidates = ['BOOLEAN', 'BIGINT', 'DECIMAL(38,10)', 'DATE', 'TIME', 'TIMESTAMP', 'VARCHAR']"
        f"{forced_types})"
    )


def _write_csv_parquet(source: Path, destination: Path) -> None:
    connection = _duckdb().connect(":memory:")
    raw_parquet = destination.with_name(f".{destination.name}.{uuid4().hex}.raw.parquet")
    try:
        header_relation = (
            f"read_csv({_sql_literal(source)}, header = true, all_varchar = true, sample_size = 1, "
            "ignore_errors = false, null_padding = false)"
        )
        description = connection.execute(f"DESCRIBE SELECT * FROM {header_relation}").fetchall()
        headers = [str(row[0]) for row in description]
        if not headers:
            raise ValueError("Import source has no headers.")
        if any(header.startswith("__aibi_") for header in headers):
            raise ValueError("Source columns cannot use the reserved __aibi_ prefix.")
        raw_relation = (
            f"read_csv({_sql_literal(source)}, header = true, all_varchar = true, "
            "ignore_errors = false, null_padding = false)"
        )
        connection.execute(
            f"""
            COPY (
              SELECT CAST(row_number() OVER () AS BIGINT) AS {_quote_identifier(INTERNAL_ROW_ID)}, *
              FROM {raw_relation}
            ) TO {_sql_literal(raw_parquet)}
            (FORMAT PARQUET, COMPRESSION ZSTD, ROW_GROUP_SIZE 122880)
            """
        )

        sample_relation = _csv_relation(source, [header for header in headers if _identifier_field(header)])
        sampled_types = {
            str(row[0]): str(row[1]).upper()
            for row in connection.execute(f"DESCRIBE SELECT * FROM {sample_relation}").fetchall()
        }
        candidates_by_field: dict[str, list[str]] = {}
        for header in headers:
            sampled_type = sampled_types.get(header, "VARCHAR")
            if _identifier_field(header):
                candidates_by_field[header] = []
            elif _measure_field(header):
                candidates_by_field[header] = ["DECIMAL(38,10)"]
            elif sampled_type == "BIGINT":
                candidates_by_field[header] = ["BIGINT", "DECIMAL(38,10)"]
            elif sampled_type == "DATE":
                candidates_by_field[header] = ["DATE", "TIMESTAMP"]
            elif sampled_type in {"BOOLEAN", "DECIMAL(38,10)", "TIME", "TIMESTAMP"}:
                candidates_by_field[header] = [sampled_type]
            else:
                candidates_by_field[header] = []

        validation_candidates = [
            (field, candidate)
            for field in headers
            for candidate in candidates_by_field[field]
        ]
        invalid_by_candidate: dict[tuple[str, str], int] = {}
        if validation_candidates:
            validation_sql: list[str] = []
            for field, candidate in validation_candidates:
                raw_value = f"trim(CAST(source.{_quote_identifier(field)} AS VARCHAR))"
                cast_value = f"replace({raw_value}, ',', '')" if candidate.startswith("DECIMAL") else raw_value
                validation_sql.append(
                    f"SUM(CASE WHEN source.{_quote_identifier(field)} IS NOT NULL "
                    f"AND {raw_value} <> '' AND TRY_CAST({cast_value} AS {candidate}) IS NULL "
                    "THEN 1 ELSE 0 END)::BIGINT"
                )
            invalid_counts = connection.execute(
                f"SELECT {', '.join(validation_sql)} FROM read_parquet({_sql_literal(raw_parquet)}) AS source"
            ).fetchone()
            invalid_by_candidate = {
                candidate: int(invalid or 0)
                for candidate, invalid in zip(validation_candidates, invalid_counts, strict=True)
            }

        selected_types = {
            field: next(
                (
                    candidate
                    for candidate in candidates_by_field[field]
                    if invalid_by_candidate.get((field, candidate), 1) == 0
                ),
                "VARCHAR",
            )
            for field in headers
        }
        public_projection = ", ".join(
            (
                f"TRY_CAST(replace(trim(CAST(source.{_quote_identifier(header)} AS VARCHAR)), ',', '') "
                f"AS DECIMAL(38,10)) AS {_quote_identifier(header)}"
                if selected_types[header].startswith("DECIMAL")
                else f"TRY_CAST(source.{_quote_identifier(header)} AS {selected_types[header]}) AS {_quote_identifier(header)}"
                if selected_types[header] != "VARCHAR"
                else f"source.{_quote_identifier(header)}"
            )
            for header in headers
        )
        connection.execute(
            f"""
            COPY (
              SELECT source.{_quote_identifier(INTERNAL_ROW_ID)}, {public_projection}
              FROM read_parquet({_sql_literal(raw_parquet)}) AS source
            ) TO {_sql_literal(destination)}
            (FORMAT PARQUET, COMPRESSION ZSTD, ROW_GROUP_SIZE 122880)
            """
        )
    finally:
        connection.close()
        raw_parquet.unlink(missing_ok=True)


def _write_parquet_stage(source: Path, destination: Path) -> None:
    """Normalize connector-produced Parquet into the stable stage row-id contract."""

    relation = f"read_parquet({_sql_literal(source)})"
    connection = _duckdb().connect(":memory:")
    try:
        description = connection.execute(f"DESCRIBE SELECT * FROM {relation}").fetchall()
        fields = [(str(row[0]), str(row[1]).upper()) for row in description]
        internal = [field for field, _data_type in fields if field.startswith("__aibi_")]
        if any(field != INTERNAL_ROW_ID for field in internal):
            raise ValueError("Connector Parquet contains unsupported internal columns.")
        public = [field for field, _data_type in fields if not field.startswith("__aibi_")]
        if not public:
            raise ValueError("Connector Parquet has no public fields.")
        public_sql = ", ".join(f"source.{_quote_identifier(field)}" for field in public)
        order_sql = (
            f"ORDER BY source.{_quote_identifier(INTERNAL_ROW_ID)}"
            if INTERNAL_ROW_ID in internal
            else ""
        )
        connection.execute(
            f"""
            COPY (
              SELECT CAST(row_number() OVER ({order_sql}) AS BIGINT) AS {_quote_identifier(INTERNAL_ROW_ID)},
                     {public_sql}
              FROM {relation} AS source
            ) TO {_sql_literal(destination)}
            (FORMAT PARQUET, COMPRESSION ZSTD, ROW_GROUP_SIZE 122880)
            """
        )
    finally:
        connection.close()


def _promote_type(current: str | None, value: Any) -> str:
    if value is None or value == "":
        return current or "UNKNOWN"
    if isinstance(value, bool):
        incoming = "BOOLEAN"
    elif isinstance(value, datetime):
        incoming = "TIMESTAMP"
    elif isinstance(value, date):
        incoming = "DATE"
    elif isinstance(value, int):
        incoming = "BIGINT"
    elif isinstance(value, float):
        incoming = "DECIMAL(38,10)"
    else:
        incoming = "VARCHAR"
    previous = current or "UNKNOWN"
    if previous == "UNKNOWN" or previous == incoming:
        return incoming
    if {previous, incoming}.issubset({"BIGINT", "DECIMAL(38,10)"}):
        return "DECIMAL(38,10)"
    return "VARCHAR"


def _xlsx_headers_and_types(source: Path) -> tuple[list[str], list[str]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as error:  # pragma: no cover - deployment contract
        raise RuntimeError("openpyxl is required for Excel import staging.") from error
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        iterator = sheet.iter_rows(values_only=True)
        try:
            headers = _normalize_headers(next(iterator))
        except StopIteration as error:
            raise ValueError("Import source has no headers.") from error
        inferred: list[str | None] = [None for _ in headers]
        for row in iterator:
            for index in range(len(headers)):
                inferred[index] = _promote_type(inferred[index], row[index] if index < len(row) else None)
        return headers, [
            "VARCHAR" if _identifier_field(header) else value if value and value != "UNKNOWN" else "VARCHAR"
            for header, value in zip(headers, inferred)
        ]
    finally:
        workbook.close()


def _cell_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value)


def _write_xlsx_parquet(source: Path, destination: Path) -> None:
    try:
        import pandas as pd  # type: ignore
        from openpyxl import load_workbook  # type: ignore
    except ImportError as error:  # pragma: no cover - deployment contract
        raise RuntimeError("pandas and openpyxl are required for Excel import staging.") from error
    headers, types = _xlsx_headers_and_types(source)
    connection = _duckdb().connect(":memory:")
    columns_sql = ", ".join(
        [f"{_quote_identifier(INTERNAL_ROW_ID)} BIGINT", *[
            f"{_quote_identifier(header)} {data_type}" for header, data_type in zip(headers, types)
        ]]
    )
    connection.execute(f"CREATE TEMP TABLE __aibi_xlsx_stage ({columns_sql})")
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        iterator = sheet.iter_rows(values_only=True)
        next(iterator, None)
        batch: list[list[Any]] = []
        row_id = 0

        def flush() -> None:
            if not batch:
                return
            frame = pd.DataFrame(batch, columns=[INTERNAL_ROW_ID, *headers], dtype="object")
            connection.register("__aibi_xlsx_batch", frame)
            try:
                select_sql = ", ".join([
                    f"CAST({_quote_identifier(INTERNAL_ROW_ID)} AS BIGINT)",
                    *[
                        f"CAST({_quote_identifier(header)} AS {data_type})"
                        for header, data_type in zip(headers, types)
                    ],
                ])
                connection.execute(f"INSERT INTO __aibi_xlsx_stage SELECT {select_sql} FROM __aibi_xlsx_batch")
            finally:
                connection.unregister("__aibi_xlsx_batch")
            batch.clear()

        for raw in iterator:
            row_id += 1
            values = [_cell_text(raw[index] if index < len(raw) else None) for index in range(len(headers))]
            batch.append([row_id, *values])
            if len(batch) >= 20_000:
                flush()
        flush()
        connection.execute(
            f"""
            COPY (SELECT * FROM __aibi_xlsx_stage ORDER BY {_quote_identifier(INTERNAL_ROW_ID)})
            TO {_sql_literal(destination)}
            (FORMAT PARQUET, COMPRESSION ZSTD, ROW_GROUP_SIZE 122880)
            """
        )
    finally:
        workbook.close()
        connection.close()


def _write_bounded_rows_parquet(
    headers: Sequence[str],
    rows: Iterable[Mapping[str, Any]],
    destination: Path,
) -> None:
    """Small-test adapter only; production callers must stage directly from a file."""

    try:
        import pandas as pd  # type: ignore
    except ImportError as error:  # pragma: no cover - deployment contract
        raise RuntimeError("pandas is required for the bounded staging helper.") from error
    normalized = _normalize_headers(headers)
    materialized: list[dict[str, Any]] = []
    for index, row in enumerate(rows, start=1):
        if index > MAX_TEST_HELPER_ROWS:
            raise ValueError("Iterable import staging is limited to bounded tests; use a source file for production imports.")
        materialized.append({header: row.get(header) for header in normalized})
    frame = pd.DataFrame(materialized, columns=normalized, dtype="object")
    for header in normalized:
        frame[header] = frame[header].map(lambda value: None if value is None else str(value))
    connection = _duckdb().connect(":memory:")
    connection.register("__aibi_test_rows", frame)
    try:
        public_columns = ", ".join(_quote_identifier(header) for header in normalized)
        connection.execute(
            f"""
            COPY (
              SELECT CAST(row_number() OVER () AS BIGINT) AS {_quote_identifier(INTERNAL_ROW_ID)}, {public_columns}
              FROM __aibi_test_rows
            ) TO {_sql_literal(destination)}
            (FORMAT PARQUET, COMPRESSION ZSTD, ROW_GROUP_SIZE 122880)
            """
        )
    finally:
        connection.unregister("__aibi_test_rows")
        connection.close()


def _infer_role(field: str, data_type: str) -> tuple[str, str, float]:
    lowered = field.casefold()
    upper_type = data_type.upper()
    if any(token in lowered for token in ("time", "date", "日期", "时间")) or upper_type.startswith(("DATE", "TIMESTAMP")):
        return "event_time", "filterable", 0.92
    if _identifier_field(field):
        return "identity_key", "joinable", 0.88
    if "status" in lowered or "状态" in field:
        return "status", "filterable", 0.84
    if upper_type.startswith(("TINYINT", "SMALLINT", "INTEGER", "BIGINT", "HUGEINT", "DECIMAL", "FLOAT", "DOUBLE", "REAL")):
        return "measure", "aggregatable", 0.9
    return "dimension", "groupable", 0.78


def profile_parquet(parquet_path: str | Path) -> dict[str, Any]:
    path = Path(parquet_path).resolve()
    relation = f"read_parquet({_sql_literal(path)})"
    connection = _duckdb().connect(":memory:")
    try:
        description = connection.execute(f"DESCRIBE SELECT * FROM {relation}").fetchall()
        public = [(str(row[0]), str(row[1]).upper()) for row in description if not str(row[0]).startswith("__aibi_")]
        if not public:
            raise ValueError("Staged Parquet has no public fields.")
        aggregate_parts = ["COUNT(*)::BIGINT AS row_count"]
        for index, (field, _data_type) in enumerate(public):
            column = _quote_identifier(field)
            aggregate_parts.append(
                f"SUM(CASE WHEN {column} IS NOT NULL AND trim(CAST({column} AS VARCHAR)) <> '' THEN 1 ELSE 0 END)::BIGINT AS nonempty_{index}"
            )
            aggregate_parts.append(f"approx_count_distinct({column})::BIGINT AS distinct_{index}")
        aggregate = connection.execute(f"SELECT {', '.join(aggregate_parts)} FROM {relation}").fetchone()
        row_count = int(aggregate[0] or 0)
        sample_rows = connection.execute(
            f"SELECT {', '.join(_quote_identifier(field) for field, _ in public)} FROM {relation} "
            f"ORDER BY {_quote_identifier(INTERNAL_ROW_ID)} LIMIT 5"
        ).fetchall()
        fields: list[dict[str, Any]] = []
        dimensions: list[str] = []
        measures: list[str] = []
        identity_keys: list[str] = []
        for index, (field, data_type) in enumerate(public):
            role, usage, confidence = _infer_role(field, data_type)
            if role == "measure":
                measures.append(field)
            elif role == "identity_key":
                identity_keys.append(field)
            elif role in {"dimension", "status"}:
                dimensions.append(field)
            samples = [str(row[index]) for row in sample_rows if row[index] is not None][:5]
            fields.append({
                "field": field,
                "role": role,
                "usage": usage,
                "confidence": confidence,
                "inferredType": data_type,
                "nonEmpty": int(aggregate[1 + index * 2] or 0),
                "uniqueCount": int(aggregate[2 + index * 2] or 0),
                "uniqueCountMode": "approximate",
                "sampleValues": samples,
                "sample": samples,
            })
        return {
            "rowCount": row_count,
            "columnCount": len(public),
            "dimensions": dimensions,
            "measures": measures,
            "identityKeys": identity_keys,
            "warnings": [] if row_count else ["Source has no data rows."],
            "fields": fields,
            "schemaFields": [{"name": field, "type": data_type} for field, data_type in public],
            "internalColumns": list(INTERNAL_SCHEMA_FIELDS),
            "profilingMode": "duckdb-single-pass-approximate",
            "sample": [
                {field: (None if row[index] is None else str(row[index])) for index, (field, _type) in enumerate(public)}
                for row in sample_rows
            ],
        }
    finally:
        connection.close()


def _manifest_material(manifest: Mapping[str, Any]) -> dict[str, Any]:
    return {key: value for key, value in manifest.items() if key != "contentFingerprint"}


def _stage_summary(manifest: Mapping[str, Any]) -> dict[str, Any]:
    return {
        key: manifest[key]
        for key in (
            "schema", "stageKey", "workspaceId", "contentHash", "contentFingerprint",
            "parserVersion", "sourceName", "sourceBytes", "rowCount", "columnCount",
            "schemaFields", "internalColumns", "schemaFingerprint", "objectHash", "objectBytes",
            "createdAt", "expiresAt", "sealed",
        )
    }


def create_import_stage(
    *,
    source_path: str | Path,
    workspace_id: str,
    headers: Sequence[str] | None = None,
    rows: Iterable[Mapping[str, Any]] | None = None,
    profile: dict[str, Any] | None = None,
    root: Path | None = None,
    timestamp: str | None = None,
) -> dict[str, Any]:
    """Stream a file into a sealed Parquet stage.

    ``headers``/``rows`` is deliberately capped and exists only for focused unit
    tests. Production import paths pass only ``source_path`` and ``workspace_id``.
    """

    resolved_source = Path(source_path).resolve()
    if not resolved_source.is_file() or resolved_source.is_symlink():
        raise FileNotFoundError(resolved_source)
    normalized_workspace = str(workspace_id or "").strip()
    workspace_bucket(normalized_workspace)
    stage_root = (root or configured_import_stage_root()).resolve()
    if stage_root.exists() and stage_root.is_symlink():
        raise PermissionError("Import stage root cannot be a symbolic link.")
    source_bytes = int(resolved_source.stat().st_size)
    content_hash = file_sha256(resolved_source)
    stage_key = "stage_" + canonical_fingerprint({
        "workspaceId": normalized_workspace,
        "contentHash": content_hash,
        "parserVersion": IMPORT_STAGE_PARSER_VERSION,
    })[:24]
    destination = _stage_directory(stage_root, normalized_workspace, stage_key)
    quota_lock = stage_root / ".aibi-import-stage-quota.lock"
    with workspace_mutation_lock(quota_lock, timeout_seconds=60.0):
        if destination.is_dir():
            manifest = load_import_stage_manifest(
                stage_key=stage_key,
                workspace_id=normalized_workspace,
                root=stage_root,
                allow_expired=True,
            )
            if _parse_iso(str(manifest["expiresAt"])) > datetime.now(timezone.utc):
                return _stage_summary(manifest)
            shutil.rmtree(destination)

        quota = _bounded_int(
            os.environ.get("AIBI_IMPORT_STAGE_MAX_BYTES"),
            DEFAULT_STAGE_QUOTA_BYTES,
            1024 * 1024,
            500 * 1024 * 1024 * 1024,
        )
        current_usage = _root_usage(stage_root)
        if source_bytes > quota or current_usage + source_bytes > quota:
            raise OSError("Import stage quota would be exceeded.")
        destination.parent.mkdir(parents=True, exist_ok=True)
        if destination.parent.is_symlink():
            raise PermissionError("Import stage workspace directory cannot be a symbolic link.")
        temporary = destination.with_name(f".{stage_key}.{uuid4().hex}.tmp")
        temporary.mkdir()
        parquet_path = temporary / PARQUET_FILE
        manifest_path = temporary / MANIFEST_FILE
        created_at = timestamp or now_iso()
        ttl_seconds = _bounded_int(
            os.environ.get("AIBI_IMPORT_STAGE_TTL_SECONDS"),
            DEFAULT_STAGE_TTL_SECONDS,
            60,
            30 * 24 * 60 * 60,
        )
        try:
            if rows is not None:
                if headers is None:
                    raise ValueError("Bounded iterable staging requires headers.")
                _write_bounded_rows_parquet(headers, rows, parquet_path)
            elif resolved_source.suffix.casefold() == ".csv":
                _write_csv_parquet(resolved_source, parquet_path)
            elif resolved_source.suffix.casefold() in {".xlsx", ".xlsm"}:
                _write_xlsx_parquet(resolved_source, parquet_path)
            elif resolved_source.suffix.casefold() == ".parquet":
                _write_parquet_stage(resolved_source, parquet_path)
            else:
                raise ValueError("Parquet staging supports CSV, XLSX, XLSM, and connector-produced Parquet sources.")

            computed_profile = profile_parquet(parquet_path)
            if profile is not None and int(profile.get("rowCount") or 0) != int(computed_profile["rowCount"]):
                raise ValueError("Import stage row count does not match its supplied test profile.")
            object_hash = file_sha256(parquet_path)
            schema_fields = list(computed_profile["schemaFields"])
            manifest: dict[str, Any] = {
                "schema": IMPORT_STAGE_SCHEMA,
                "stageKey": stage_key,
                "workspaceId": normalized_workspace,
                "contentHash": content_hash,
                "parserVersion": IMPORT_STAGE_PARSER_VERSION,
                "sourceName": resolved_source.name,
                "sourceBytes": source_bytes,
                "headers": [field["name"] for field in schema_fields],
                "rowCount": int(computed_profile["rowCount"]),
                "columnCount": len(schema_fields),
                "schemaFields": schema_fields,
                "internalColumns": list(INTERNAL_SCHEMA_FIELDS),
                "schemaFingerprint": canonical_fingerprint({"fields": schema_fields}),
                "objectHash": object_hash,
                "objectBytes": int(parquet_path.stat().st_size),
                "profile": computed_profile,
                "createdAt": created_at,
                "expiresAt": _expires_at(created_at, ttl_seconds),
                "sealed": True,
            }
            manifest["contentFingerprint"] = canonical_fingerprint(_manifest_material(manifest))
            manifest_path.write_text(
                json.dumps(manifest, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=str),
                encoding="utf-8",
            )
            _fsync_file(parquet_path)
            _fsync_file(manifest_path)
            stage_bytes = sum(path.stat().st_size for path in temporary.iterdir() if path.is_file())
            if stage_bytes > quota or current_usage + stage_bytes > quota:
                raise OSError("Import stage quota would be exceeded by the parsed stage.")
            os.replace(temporary, destination)
            _fsync_directory(destination.parent)
        finally:
            if temporary.exists():
                shutil.rmtree(temporary)
        return _stage_summary(manifest)


def load_import_stage_manifest(
    *,
    stage_key: str,
    workspace_id: str,
    root: Path | None = None,
    allow_expired: bool = True,
) -> dict[str, Any]:
    stage_root = (root or configured_import_stage_root()).resolve()
    parquet_path, manifest_path = _stage_paths(stage_root, workspace_id, stage_key)
    directory = parquet_path.parent
    if directory.is_symlink() or parquet_path.is_symlink() or manifest_path.is_symlink():
        raise PermissionError("Import stage files cannot be symbolic links.")
    if not parquet_path.is_file() or not manifest_path.is_file():
        raise FileNotFoundError(f"Import stage is unavailable: {stage_key}")
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    if manifest.get("schema") != IMPORT_STAGE_SCHEMA or manifest.get("parserVersion") != IMPORT_STAGE_PARSER_VERSION:
        raise ValueError("Import stage schema is unsupported.")
    if manifest.get("stageKey") != stage_key or manifest.get("workspaceId") != workspace_id:
        raise PermissionError("Import stage does not belong to the requested workspace.")
    if manifest.get("sealed") is not True:
        raise ValueError("Import stage is not sealed.")
    if canonical_fingerprint(_manifest_material(manifest)) != manifest.get("contentFingerprint"):
        raise ValueError("Import stage manifest integrity verification failed.")
    if file_sha256(parquet_path) != str(manifest.get("objectHash") or ""):
        raise ValueError("Import stage object integrity verification failed.")
    if int(parquet_path.stat().st_size) != int(manifest.get("objectBytes") or -1):
        raise ValueError("Import stage object size drifted.")
    if list(manifest.get("internalColumns") or []) != list(INTERNAL_SCHEMA_FIELDS):
        raise ValueError("Import stage internal row id contract is invalid.")
    if not allow_expired and _parse_iso(str(manifest["expiresAt"])) <= datetime.now(timezone.utc):
        raise ValueError("Import stage expired; run the preview again.")
    return manifest


def resolve_import_stage_parquet(
    *,
    stage_key: str,
    workspace_id: str,
    root: Path | None = None,
    allow_expired: bool = True,
) -> tuple[Path, dict[str, Any]]:
    manifest = load_import_stage_manifest(
        stage_key=stage_key,
        workspace_id=workspace_id,
        root=root,
        allow_expired=allow_expired,
    )
    stage_root = (root or configured_import_stage_root()).resolve()
    parquet_path, _manifest_path = _stage_paths(stage_root, workspace_id, stage_key)
    return parquet_path, manifest


def iter_import_stage_rows(
    *,
    stage_key: str,
    workspace_id: str,
    root: Path | None = None,
    batch_size: int = 1000,
) -> Iterator[list[dict[str, Any]]]:
    parquet_path, manifest = resolve_import_stage_parquet(
        stage_key=stage_key,
        workspace_id=workspace_id,
        root=root,
    )
    if int(manifest["rowCount"]) > MAX_TEST_HELPER_ROWS:
        raise ValueError("Materialized stage reads are limited to bounded tests; query the Parquet stage with DuckDB.")
    headers = [str(header) for header in manifest["headers"]]
    connection = _duckdb().connect(":memory:")
    try:
        cursor = connection.execute(
            f"SELECT {', '.join(_quote_identifier(header) for header in headers)} "
            f"FROM read_parquet({_sql_literal(parquet_path)}) ORDER BY {_quote_identifier(INTERNAL_ROW_ID)}"
        )
        while records := cursor.fetchmany(max(1, min(int(batch_size), 10_000))):
            yield [
                {header: ("" if row[index] is None else str(row[index])) for index, header in enumerate(headers)}
                for row in records
            ]
    finally:
        connection.close()


def read_import_stage(
    *,
    stage_key: str,
    workspace_id: str,
    root: Path | None = None,
) -> tuple[list[str], list[dict[str, Any]], dict[str, Any], dict[str, Any]]:
    manifest = load_import_stage_manifest(stage_key=stage_key, workspace_id=workspace_id, root=root)
    rows = [row for batch in iter_import_stage_rows(stage_key=stage_key, workspace_id=workspace_id, root=root) for row in batch]
    return list(manifest["headers"]), rows, dict(manifest["profile"]), _stage_summary(manifest)


def validate_import_stage_for_confirmation(
    *,
    stage_key: str,
    workspace_id: str,
    root: Path | None = None,
) -> dict[str, Any]:
    manifest = load_import_stage_manifest(
        stage_key=stage_key,
        workspace_id=workspace_id,
        root=root,
        allow_expired=False,
    )
    return _stage_summary(manifest)
