from __future__ import annotations

import csv
import hashlib
import json
import os
import re
import sqlite3
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Callable
from urllib.error import HTTPError
from urllib.parse import parse_qsl, quote, urlencode, urljoin, urlsplit, urlunsplit
from urllib.request import HTTPRedirectHandler, Request, build_opener

from openpyxl import load_workbook

from bi_cli_core import ROOT, source_label


ADAPTER_SCHEMA = "aibi-connector-adapter/v1"
ADAPTER_RECEIPT_SCHEMA = "aibi-connector-adapter-receipt/v1"
LOCAL_TABULAR_ADAPTER = "local-tabular/v1"
HTTP_JSON_ADAPTER = "http-json/v1"
SQLITE_TABLE_ADAPTER = "sqlite-table/v1"
MAX_SOURCE_BYTES = 100 * 1024 * 1024
MAX_HTTP_BYTES = 5 * 1024 * 1024
MAX_HTTP_PAGES = 10
MAX_ADAPTER_ROWS = 10_000
HTTP_TIMEOUT_SECONDS = 10
MAX_PREVIEW_ROWS = 100
MAX_PREVIEW_COLUMNS = 64
MAX_CELL_TEXT = 512
ALLOWED_FILE_SUFFIXES = {".csv", ".xlsx", ".xlsm"}
FORBIDDEN_REPOSITORY_PARTS = {
    "aibi-a",
    "aibi-b",
    "aibi-d",
    "aibi-e",
    "aibi项目杂交",
}


def _canonical(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=str)


def _fingerprint(value: Any) -> str:
    return hashlib.sha256(_canonical(value).encode("utf-8")).hexdigest()


def adapter_contracts() -> list[dict[str, Any]]:
    local_contract = {
        "schema": ADAPTER_SCHEMA,
        "adapterId": LOCAL_TABULAR_ADAPTER,
        "connectorType": "file",
        "available": True,
        "operations": ["metadata-discovery", "bounded-preview", "sync-plan"],
        "permissions": {
            "filesystem": "selected-file-read-only",
            "network": "none",
            "database": "metadata-read-only",
            "arbitrarySql": False,
            "otherAibiRepositories": "forbidden",
        },
        "limits": {
            "sourceBytes": MAX_SOURCE_BYTES,
            "previewRows": MAX_PREVIEW_ROWS,
            "previewColumns": MAX_PREVIEW_COLUMNS,
            "extensions": sorted(ALLOWED_FILE_SUFFIXES),
        },
        "credentials": {"mode": "none", "valuesInReceipts": False},
        "persistence": {"performedByAdapter": False, "boundary": "sync-connector --yes"},
    }
    http_contract = {
        "schema": ADAPTER_SCHEMA,
        "adapterId": HTTP_JSON_ADAPTER,
        "connectorType": "api",
        "available": True,
        "operations": ["metadata-discovery", "bounded-preview", "sync-plan"],
        "permissions": {
            "filesystem": "none", "network": "allowlisted-get-only", "database": "metadata-read-only",
            "arbitrarySql": False, "otherAibiRepositories": "forbidden",
        },
        "limits": {"responseBytes": MAX_HTTP_BYTES, "pages": MAX_HTTP_PAGES, "rows": MAX_ADAPTER_ROWS, "timeoutSeconds": HTTP_TIMEOUT_SECONDS},
        "credentials": {"mode": "server-env-bearer-reference", "valuesInReceipts": False},
        "persistence": {"performedByAdapter": False, "boundary": "sync-connector --yes"},
    }
    sqlite_contract = {
        "schema": ADAPTER_SCHEMA,
        "adapterId": SQLITE_TABLE_ADAPTER,
        "connectorType": "database",
        "available": True,
        "operations": ["metadata-discovery", "bounded-preview", "sync-plan"],
        "permissions": {
            "filesystem": "allowlisted-sqlite-read-only", "network": "none", "database": "selected-table-read-only",
            "arbitrarySql": False, "otherAibiRepositories": "forbidden",
        },
        "limits": {"sourceBytes": MAX_SOURCE_BYTES, "rows": MAX_ADAPTER_ROWS, "previewRows": MAX_PREVIEW_ROWS, "previewColumns": MAX_PREVIEW_COLUMNS},
        "credentials": {"mode": "none", "valuesInReceipts": False},
        "persistence": {"performedByAdapter": False, "boundary": "sync-connector --yes"},
    }
    unavailable = []
    for connector_type in ("erp",):
        unavailable.append({
            "schema": ADAPTER_SCHEMA,
            "adapterId": f"{connector_type}-allowlist/unavailable",
            "connectorType": connector_type,
            "available": False,
            "operations": [],
            "permissions": {
                "filesystem": "none",
                "network": "none",
                "database": "none",
                "arbitrarySql": False,
                "otherAibiRepositories": "forbidden",
            },
            "credentials": {"mode": "server-env-reference", "valuesInReceipts": False},
            "reason": "No provider-specific allowlist adapter is installed; registration grants no execution authority.",
        })
    return [local_contract, http_contract, sqlite_contract, *unavailable]


def adapter_contract_for(connector_type: str) -> dict[str, Any]:
    contract = next((item for item in adapter_contracts() if item["connectorType"] == connector_type), None)
    if not contract:
        raise ValueError(f"No adapter contract exists for connector type: {connector_type}")
    return contract


def validate_credential_reference(value: str | None) -> str:
    reference = str(value or "").strip()
    if not reference:
        return ""
    if not reference.startswith("env:"):
        raise ValueError("Connector credentials must use a server-side env:NAME reference; literal values are forbidden.")
    name = reference[4:]
    if not name or not (name[0].isalpha() or name[0] == "_") or not all(char.isalnum() or char == "_" for char in name):
        raise ValueError("Connector credential reference must match env:NAME.")
    return reference


def _candidate_path(value: str) -> Path:
    raw = Path(str(value or "").strip())
    if not str(raw):
        raise ValueError("Connector endpoint is required for the local tabular adapter.")
    return raw if raw.is_absolute() else ROOT / raw


def _reject_symlink_chain(candidate: Path) -> None:
    current = candidate
    while True:
        if current.exists() and (current.is_symlink() or (hasattr(current, "is_junction") and current.is_junction())):
            raise ValueError("Connector resources cannot traverse symbolic links or junction-like links.")
        if current.parent == current:
            break
        current = current.parent


def validate_connector_resource(value: str, *, require_exists: bool = True) -> Path:
    candidate = _candidate_path(value)
    _reject_symlink_chain(candidate)
    resolved = candidate.resolve()
    lowered_parts = {part.casefold() for part in resolved.parts}
    forbidden = sorted(lowered_parts.intersection(FORBIDDEN_REPOSITORY_PARTS))
    if forbidden:
        raise ValueError(f"Connector resource belongs to a forbidden AIBI repository boundary: {forbidden[0]}")
    if resolved.suffix.casefold() not in ALLOWED_FILE_SUFFIXES:
        raise ValueError(f"Local tabular adapter accepts only: {', '.join(sorted(ALLOWED_FILE_SUFFIXES))}")
    if require_exists and not resolved.is_file():
        raise FileNotFoundError(resolved.name)
    if resolved.exists() and resolved.stat().st_size > MAX_SOURCE_BYTES:
        raise ValueError(f"Connector resource exceeds the {MAX_SOURCE_BYTES} byte adapter limit.")
    return resolved


def _file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _safe_scalar(value: Any) -> str | int | float | bool | None:
    if value is None or isinstance(value, (bool, int, float)):
        return value
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    text = str(value)
    return text if len(text) <= MAX_CELL_TEXT else f"{text[:MAX_CELL_TEXT]}…"


def _csv_preview(path: Path, limit: int) -> tuple[list[str], list[list[Any]], bool, str]:
    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                reader = csv.reader(handle)
                headers = [str(item).strip() or f"column_{index + 1}" for index, item in enumerate(next(reader, []))]
                rows: list[list[Any]] = []
                for row in reader:
                    rows.append([_safe_scalar(item) for item in row[:MAX_PREVIEW_COLUMNS]])
                    if len(rows) > limit:
                        break
                return headers[:MAX_PREVIEW_COLUMNS], rows[:limit], len(rows) > limit, "csv"
        except UnicodeDecodeError as error:
            last_error = error
    raise ValueError(f"Connector CSV encoding is unsupported: {last_error}")


def _workbook_preview(path: Path, limit: int) -> tuple[list[str], list[list[Any]], bool, str]:
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        worksheet = workbook.active
        iterator = worksheet.iter_rows(values_only=True)
        header_row = next(iterator, ())
        headers = [str(item).strip() if item is not None and str(item).strip() else f"column_{index + 1}" for index, item in enumerate(header_row)]
        rows: list[list[Any]] = []
        for row in iterator:
            rows.append([_safe_scalar(item) for item in row[:MAX_PREVIEW_COLUMNS]])
            if len(rows) > limit:
                break
        return headers[:MAX_PREVIEW_COLUMNS], rows[:limit], len(rows) > limit or worksheet.max_row > limit + 1, worksheet.title
    finally:
        workbook.close()


def bounded_tabular_preview(path: Path, limit: int) -> dict[str, Any]:
    bounded_limit = max(1, min(int(limit), MAX_PREVIEW_ROWS))
    if path.suffix.casefold() == ".csv":
        columns, rows, truncated, source_part = _csv_preview(path, bounded_limit)
    else:
        columns, rows, truncated, source_part = _workbook_preview(path, bounded_limit)
    normalized_rows = []
    for row in rows:
        normalized_rows.append({column: row[index] if index < len(row) else None for index, column in enumerate(columns)})
    return {
        "columns": columns,
        "rows": normalized_rows,
        "returnedRows": len(normalized_rows),
        "truncated": truncated,
        "sourcePart": source_part,
        "limits": {"rows": bounded_limit, "columns": MAX_PREVIEW_COLUMNS, "cellText": MAX_CELL_TEXT},
    }


def public_connector(row: Any) -> dict[str, Any]:
    raw_config = json.loads(row["config_json"] or "{}")
    public_config = {key: value for key, value in raw_config.items() if key != "credentialRef"}
    public_config["credentialConfigured"] = bool(raw_config.get("credentialRef"))
    return {
        "connectorKey": row["connector_key"],
        "workspaceId": row["workspace_id"],
        "name": row["name"],
        "type": row["connector_type"],
        "provider": row["provider"],
        "status": row["status"],
        "config": public_config,
        "schedule": json.loads(row["schedule_json"] or "{}"),
        "createdAt": row["created_at"],
        "updatedAt": row["updated_at"],
        "lastSyncAt": row["last_sync_at"],
        "lastSyncStatus": row["last_sync_status"],
        "lastSyncResult": json.loads(row["last_sync_result_json"] or "null"),
    }


def _connector_identity(row: Any) -> dict[str, Any]:
    raw_config = json.loads(row["config_json"] or "{}")
    return {
        "connectorKey": row["connector_key"],
        "workspaceId": row["workspace_id"],
        "name": row["name"],
        "type": row["connector_type"],
        "provider": row["provider"],
        "status": row["status"],
        "targetTableKey": str(raw_config.get("targetTableKey") or ""),
        "credentialConfigured": bool(raw_config.get("credentialRef")),
    }


def _blocked_receipt(row: Any, operation: str) -> dict[str, Any]:
    identity = _connector_identity(row)
    contract = adapter_contract_for(identity["type"])
    return {
        "ok": False,
        "blocked": True,
        "schema": ADAPTER_RECEIPT_SCHEMA,
        "operation": operation,
        "connector": identity,
        "adapter": contract,
        "reason": contract.get("reason") or "Connector adapter is unavailable.",
        "sideEffects": {"network": False, "businessDatabaseWrite": False, "artifactWrite": False},
        "evidence": ["connector-adapter-contract", "aibi-no-external-execution-without-provider-allowlist"],
    }


def _local_context(row: Any, preview_limit: int) -> tuple[dict[str, Any], Path, dict[str, Any], dict[str, Any]]:
    identity = _connector_identity(row)
    if identity["type"] != "file":
        raise ValueError("Only file connectors can use the local tabular adapter.")
    raw_config = json.loads(row["config_json"] or "{}")
    if raw_config.get("credentialRef"):
        raise ValueError("Local file connectors cannot declare credentials.")
    path = validate_connector_resource(str(raw_config.get("endpoint") or ""))
    digest = _file_sha256(path)
    resource = {
        "label": path.name,
        "extension": path.suffix.casefold(),
        "bytes": path.stat().st_size,
        "sha256": digest,
    }
    preview = bounded_tabular_preview(path, preview_limit)
    return identity, path, resource, preview


def _allowlist_values(name: str) -> list[str]:
    return [item.strip() for item in str(os.environ.get(name) or "").split(",") if item.strip()]


def _http_origin(url: str) -> str:
    parsed = urlsplit(url)
    if not parsed.scheme or not parsed.hostname or parsed.username or parsed.password:
        raise ValueError("HTTP connector endpoint must be an absolute URL without embedded credentials.")
    port = parsed.port
    default_port = 443 if parsed.scheme == "https" else 80
    authority = parsed.hostname.casefold() + (f":{port}" if port and port != default_port else "")
    return f"{parsed.scheme.casefold()}://{authority}"


def _validate_http_endpoint(url: str) -> str:
    origin = _http_origin(url)
    parsed = urlsplit(url)
    loopback = parsed.hostname in {"127.0.0.1", "localhost", "::1"}
    if parsed.scheme.casefold() != "https" and not (parsed.scheme.casefold() == "http" and loopback):
        raise ValueError("HTTP connector endpoints require HTTPS; loopback HTTP is allowed only for local verification.")
    sensitive_query_keys = {
        key for key, _value in parse_qsl(parsed.query, keep_blank_values=True)
        if any(token in key.casefold() for token in ("token", "secret", "password", "api_key", "apikey", "credential"))
    }
    if sensitive_query_keys:
        raise ValueError("HTTP connector credentials cannot be embedded in endpoint query parameters; use env:NAME.")
    allowed = {_http_origin(item) for item in _allowlist_values("AIBI_HTTP_CONNECTOR_ALLOWLIST")}
    if origin not in allowed:
        raise ValueError(f"HTTP connector origin is not allowlisted: {origin}")
    return origin


def validate_http_connector_registration(endpoint: str) -> str:
    """Validate a saved HTTP connector without issuing a network request."""
    return _validate_http_endpoint(str(endpoint or "").strip())


class _AllowlistRedirectHandler(HTTPRedirectHandler):
    def redirect_request(self, req, fp, code, msg, headers, newurl):  # noqa: ANN001
        _validate_http_endpoint(urljoin(req.full_url, newurl))
        return super().redirect_request(req, fp, code, msg, headers, newurl)


def _json_path(value: Any, path: str) -> Any:
    current = value
    for part in [item for item in path.split(".") if item]:
        if not isinstance(current, dict) or part not in current:
            raise ValueError(f"HTTP connector resource path was not found: {path}")
        current = current[part]
    return current


def _normalize_object_rows(value: Any) -> list[dict[str, Any]]:
    items = value if isinstance(value, list) else [value]
    rows: list[dict[str, Any]] = []
    for item in items:
        if isinstance(item, dict):
            row = {str(key): _safe_scalar(cell) for key, cell in item.items() if not isinstance(cell, (dict, list, tuple, set))}
        elif isinstance(item, (str, int, float, bool)) or item is None:
            row = {"value": _safe_scalar(item)}
        else:
            continue
        rows.append(dict(list(row.items())[:MAX_PREVIEW_COLUMNS]))
    return rows


def _http_page_url(endpoint: str, page: int, page_size: int, page_param: str, page_size_param: str) -> str:
    if page == 1 and not page_param:
        return endpoint
    parsed = urlsplit(endpoint)
    query = dict(parse_qsl(parsed.query, keep_blank_values=True))
    if page_param:
        query[page_param] = str(page)
    if page_size_param:
        query[page_size_param] = str(page_size)
    return urlunsplit((parsed.scheme, parsed.netloc, parsed.path, urlencode(query), parsed.fragment))


def _http_context(row: Any, row_limit: int) -> tuple[dict[str, Any], dict[str, Any], dict[str, Any]]:
    identity = _connector_identity(row)
    raw_config = json.loads(row["config_json"] or "{}")
    endpoint = str(raw_config.get("endpoint") or "").strip()
    origin = _validate_http_endpoint(endpoint)
    resource_path = str(raw_config.get("resource") or "").strip()
    page_param = str(raw_config.get("pageParam") or "").strip()
    page_size_param = str(raw_config.get("pageSizeParam") or "").strip()
    page_size = max(1, min(int(raw_config.get("pageSize") or min(row_limit, 100)), 1000))
    max_pages = max(1, min(int(raw_config.get("maxPages") or 1), MAX_HTTP_PAGES))
    credential_ref = validate_credential_reference(raw_config.get("credentialRef"))
    headers = {"Accept": "application/json", "User-Agent": "AIBI-C/http-json-v1"}
    if credential_ref:
        credential = os.environ.get(credential_ref[4:])
        if not credential:
            raise ValueError("HTTP connector credential environment variable is not configured.")
        headers["Authorization"] = f"Bearer {credential}"
    rows: list[dict[str, Any]] = []
    total_bytes = 0
    pages_read = 0
    opener = build_opener(_AllowlistRedirectHandler())
    for page in range(1, max_pages + 1):
        page_url = _http_page_url(endpoint, page, page_size, page_param, page_size_param)
        _validate_http_endpoint(page_url)
        try:
            with opener.open(Request(page_url, headers=headers, method="GET"), timeout=HTTP_TIMEOUT_SECONDS) as response:
                _validate_http_endpoint(response.geturl())
                content_type = str(response.headers.get("Content-Type") or "").casefold()
                if "json" not in content_type:
                    raise ValueError("HTTP connector response must declare a JSON content type.")
                body = response.read(MAX_HTTP_BYTES - total_bytes + 1)
        except HTTPError as error:
            raise ValueError(f"HTTP connector request failed with status {error.code}.") from error
        total_bytes += len(body)
        if total_bytes > MAX_HTTP_BYTES:
            raise ValueError(f"HTTP connector responses exceed the {MAX_HTTP_BYTES} byte limit.")
        try:
            payload = json.loads(body.decode("utf-8"))
        except (UnicodeDecodeError, json.JSONDecodeError) as error:
            raise ValueError("HTTP connector response is not valid UTF-8 JSON.") from error
        page_rows = _normalize_object_rows(_json_path(payload, resource_path) if resource_path else payload)
        rows.extend(page_rows)
        pages_read += 1
        if len(rows) >= row_limit or not page_param or len(page_rows) < page_size:
            break
    rows = rows[:row_limit]
    columns = list(dict.fromkeys(key for item in rows for key in item))[:MAX_PREVIEW_COLUMNS]
    resource = {
        "label": f"{urlsplit(endpoint).hostname}{urlsplit(endpoint).path}",
        "origin": origin,
        "bytes": total_bytes,
        "sha256": _fingerprint(rows),
        "pagesRead": pages_read,
    }
    preview = {
        "columns": columns,
        "rows": [{key: item.get(key) for key in columns} for item in rows],
        "returnedRows": len(rows),
        "truncated": len(rows) >= row_limit,
        "sourcePart": resource_path or "$",
        "limits": {"rows": row_limit, "columns": MAX_PREVIEW_COLUMNS, "cellText": MAX_CELL_TEXT, "pages": max_pages},
    }
    return identity, resource, preview


def _validate_sqlite_resource(endpoint: str) -> Path:
    candidate = _candidate_path(endpoint)
    _reject_symlink_chain(candidate)
    resolved = candidate.resolve()
    forbidden = {part.casefold() for part in resolved.parts}.intersection(FORBIDDEN_REPOSITORY_PARTS)
    if forbidden:
        raise ValueError(f"Connector resource belongs to a forbidden AIBI repository boundary: {sorted(forbidden)[0]}")
    if resolved.suffix.casefold() not in {".sqlite", ".sqlite3", ".db"} or not resolved.is_file():
        raise ValueError("SQLite connector endpoint must be an existing .sqlite, .sqlite3, or .db file.")
    allowed = [Path(item).expanduser().resolve() for item in _allowlist_values("AIBI_SQLITE_CONNECTOR_ALLOWLIST")]
    if not any(resolved == item or (item.is_dir() and resolved.is_relative_to(item)) for item in allowed):
        raise ValueError("SQLite connector endpoint is outside AIBI_SQLITE_CONNECTOR_ALLOWLIST.")
    if resolved.stat().st_size > MAX_SOURCE_BYTES:
        raise ValueError(f"SQLite connector resource exceeds the {MAX_SOURCE_BYTES} byte limit.")
    return resolved


def validate_sqlite_connector_registration(endpoint: str, resource: str) -> Path:
    """Validate the fixed database file and table surface before persistence."""
    path = _validate_sqlite_resource(str(endpoint or "").strip())
    table = str(resource or "").strip()
    if not re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]{0,127}", table) or table.casefold().startswith("sqlite_"):
        raise ValueError("SQLite connector resource must be one explicit non-system table name.")
    return path


def _sqlite_context(row: Any, row_limit: int) -> tuple[dict[str, Any], Path, dict[str, Any], dict[str, Any]]:
    identity = _connector_identity(row)
    raw_config = json.loads(row["config_json"] or "{}")
    if raw_config.get("credentialRef"):
        raise ValueError("SQLite connectors cannot declare credentials.")
    path = _validate_sqlite_resource(str(raw_config.get("endpoint") or ""))
    table = str(raw_config.get("resource") or "").strip()
    if not re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]{0,127}", table) or table.casefold().startswith("sqlite_"):
        raise ValueError("SQLite connector resource must be one explicit non-system table name.")
    quoted = '"' + table.replace('"', '""') + '"'
    connection = sqlite3.connect(f"file:{quote(path.as_posix(), safe='/:')}?mode=ro", uri=True)
    try:
        table_row = connection.execute(
            "SELECT 1 FROM sqlite_master WHERE type = 'table' AND name = ? AND name NOT LIKE 'sqlite_%'",
            (table,),
        ).fetchone()
        if not table_row:
            raise ValueError(f"SQLite connector table does not exist: {table}")
        columns = [str(item[1]) for item in connection.execute(f"PRAGMA table_info({quoted})").fetchall()][:MAX_PREVIEW_COLUMNS]
        if not columns:
            raise ValueError(f"SQLite connector table does not exist: {table}")
        fetched = connection.execute(f"SELECT * FROM {quoted} LIMIT ?", (row_limit + 1,)).fetchall()
    finally:
        connection.close()
    rows = [{column: _safe_scalar(record[index]) for index, column in enumerate(columns)} for record in fetched[:row_limit]]
    resource = {"label": path.name, "table": table, "bytes": path.stat().st_size, "sha256": _file_sha256(path)}
    preview = {
        "columns": columns,
        "rows": rows,
        "returnedRows": len(rows),
        "truncated": len(fetched) > row_limit,
        "sourcePart": table,
        "limits": {"rows": row_limit, "columns": MAX_PREVIEW_COLUMNS, "cellText": MAX_CELL_TEXT},
    }
    return identity, path, resource, preview


def discover_connector(row: Any) -> dict[str, Any]:
    identity = _connector_identity(row)
    if identity["type"] == "file":
        identity, _path, resource, preview = _local_context(row, 1)
        adapter_id = LOCAL_TABULAR_ADAPTER
        network = False
    elif identity["type"] == "api":
        identity, resource, preview = _http_context(row, 1)
        adapter_id = HTTP_JSON_ADAPTER
        network = True
    elif identity["type"] == "database":
        identity, _path, resource, preview = _sqlite_context(row, 1)
        adapter_id = SQLITE_TABLE_ADAPTER
        network = False
    else:
        return _blocked_receipt(row, "metadata-discovery")
    metadata = {
        "resource": resource,
        "columns": preview["columns"],
        "columnCount": len(preview["columns"]),
        "sourcePart": preview["sourcePart"],
        "hasDataRows": preview["returnedRows"] > 0,
    }
    return {
        "ok": True,
        "schema": ADAPTER_RECEIPT_SCHEMA,
        "operation": "metadata-discovery",
        "connector": identity,
        "adapter": adapter_contract_for(identity["type"]),
        "metadata": metadata,
        "receiptFingerprint": _fingerprint({"operation": "metadata-discovery", "connector": identity, "metadata": metadata}),
        "sideEffects": {"network": network, "businessDatabaseWrite": False, "artifactWrite": False},
        "evidence": ["connector-adapter-contract", "selected-file-metadata", "resource-fingerprint"],
    }


def preview_connector(row: Any, limit: int) -> dict[str, Any]:
    identity = _connector_identity(row)
    if identity["type"] == "file":
        identity, _path, resource, preview = _local_context(row, limit)
        adapter_id = LOCAL_TABULAR_ADAPTER
        network = False
    elif identity["type"] == "api":
        identity, resource, preview = _http_context(row, max(1, min(int(limit), MAX_PREVIEW_ROWS)))
        adapter_id = HTTP_JSON_ADAPTER
        network = True
    elif identity["type"] == "database":
        identity, _path, resource, preview = _sqlite_context(row, max(1, min(int(limit), MAX_PREVIEW_ROWS)))
        adapter_id = SQLITE_TABLE_ADAPTER
        network = False
    else:
        return _blocked_receipt(row, "bounded-preview")
    payload = {"resource": resource, "preview": preview}
    return {
        "ok": True,
        "schema": ADAPTER_RECEIPT_SCHEMA,
        "operation": "bounded-preview",
        "connector": identity,
        "adapterId": adapter_id,
        **payload,
        "receiptFingerprint": _fingerprint({"operation": "bounded-preview", "connector": identity, **payload}),
        "sideEffects": {"network": network, "businessDatabaseWrite": False, "artifactWrite": False},
        "evidence": ["connector-adapter-contract", "bounded-scalar-preview", "resource-fingerprint"],
    }


def build_sync_plan(
    row: Any,
    *,
    connection: Any,
    registry_for_table: Callable[[Any, str], Any],
    saved_import_policy: Callable[[Any, str], dict[str, Any] | None],
) -> dict[str, Any]:
    identity = _connector_identity(row)
    source_path: Path | None = None
    materialized_rows: list[dict[str, Any]] | None = None
    if identity["type"] == "file":
        identity, source_path, resource, preview = _local_context(row, 5)
        adapter_id = LOCAL_TABULAR_ADAPTER
        network = False
    elif identity["type"] == "api":
        identity, resource, preview = _http_context(row, MAX_ADAPTER_ROWS)
        adapter_id = HTTP_JSON_ADAPTER
        network = True
        materialized_rows = list(preview["rows"])
    elif identity["type"] == "database":
        identity, _database_path, resource, preview = _sqlite_context(row, MAX_ADAPTER_ROWS)
        adapter_id = SQLITE_TABLE_ADAPTER
        network = False
        materialized_rows = list(preview["rows"])
    else:
        return _blocked_receipt(row, "sync-plan")
    raw_config = json.loads(row["config_json"] or "{}")
    target_table = str(raw_config.get("targetTableKey") or "").strip() or identity["connectorKey"]
    requested_mode = str(raw_config.get("importMode") or "auto")
    registry = registry_for_table(connection, target_table)
    resolved_mode = "merge" if requested_mode == "auto" and registry else "create" if requested_mode == "auto" else requested_mode
    unique_fields = [str(item) for item in raw_config.get("uniqueFields") or []]
    conflict_rule = str(raw_config.get("conflictRule") or "overwrite")
    if resolved_mode == "merge" and not unique_fields:
        policy = saved_import_policy(connection, target_table)
        unique_fields = [str(item) for item in (policy or {}).get("uniqueFields") or []]
        conflict_rule = str((policy or {}).get("conflictRule") or conflict_rule)
    blockers: list[str] = []
    if resolved_mode == "merge" and not unique_fields:
        blockers.append("merge-requires-unique-fields")
    plan = {
        "adapterId": adapter_id,
        "connectorKey": identity["connectorKey"],
        "resource": resource,
        "sourcePart": preview["sourcePart"],
        "columns": preview["columns"],
        "targetTableKey": target_table,
        "requestedMode": requested_mode,
        "resolvedMode": resolved_mode,
        "uniqueFields": unique_fields,
        "conflictRule": conflict_rule,
        "blockers": blockers,
        "writeBoundary": {
            "adapterWrites": False,
            "command": "sync-connector",
            "requiresConfirmation": True,
            "confirmationFlag": "--yes",
        },
    }
    return {
        "ok": not blockers,
        "blocked": bool(blockers),
        "dryRun": True,
        "requiresConfirmation": not blockers,
        "schema": ADAPTER_RECEIPT_SCHEMA,
        "operation": "sync-plan",
        "connector": identity,
        "syncPlan": plan,
        "planFingerprint": _fingerprint(plan),
        "sideEffects": {"network": network, "businessDatabaseWrite": False, "artifactWrite": False},
        "evidence": ["connector-adapter-contract", "bounded-sync-plan", "aibi-action-confirmation-boundary"],
        "_sourcePath": source_path,
        "_rows": materialized_rows,
    }


def public_sync_plan(plan: dict[str, Any]) -> dict[str, Any]:
    return {key: value for key, value in plan.items() if not key.startswith("_")}


def list_connector_adapters_command(_args: Any) -> dict[str, Any]:
    contracts = adapter_contracts()
    return {"ok": True, "schema": ADAPTER_SCHEMA, "adapters": contracts, "count": len(contracts)}


def _row_for(args: Any, open_db: Callable[[], Any], connector_by_key_or_name: Callable[[Any, str], Any]):
    connection = open_db()
    row = connector_by_key_or_name(connection, args.connector)
    if not row:
        connection.close()
        raise ValueError(f"Unknown connector: {args.connector}")
    return connection, row


def discover_connector_command(args: Any, *, open_db: Callable[[], Any], connector_by_key_or_name: Callable[[Any, str], Any]) -> dict[str, Any]:
    connection, row = _row_for(args, open_db, connector_by_key_or_name)
    try:
        return discover_connector(row)
    finally:
        connection.close()


def preview_connector_command(args: Any, *, open_db: Callable[[], Any], connector_by_key_or_name: Callable[[Any, str], Any]) -> dict[str, Any]:
    connection, row = _row_for(args, open_db, connector_by_key_or_name)
    try:
        return preview_connector(row, args.limit)
    finally:
        connection.close()


def plan_connector_sync_command(
    args: Any,
    *,
    open_db: Callable[[], Any],
    connector_by_key_or_name: Callable[[Any, str], Any],
    registry_for_table: Callable[[Any, str], Any],
    saved_import_policy: Callable[[Any, str], dict[str, Any] | None],
) -> dict[str, Any]:
    connection, row = _row_for(args, open_db, connector_by_key_or_name)
    try:
        return public_sync_plan(build_sync_plan(
            row,
            connection=connection,
            registry_for_table=registry_for_table,
            saved_import_policy=saved_import_policy,
        ))
    finally:
        connection.close()
